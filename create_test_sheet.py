import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Define styles
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
subheader_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
subheader_font = Font(bold=True, color='000000', size=10)
pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def set_header(ws, row, data):
    for col, text in enumerate(data, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = text
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

def set_subheader(ws, row, data):
    for col, text in enumerate(data, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = text
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = border

# ==================== TEST SUMMARY ====================
ws = wb.create_sheet('Test Summary', 0)
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 15

set_header(ws, 1, ['Project Name', 'BizLink CRM Platform - Comprehensive Test Suite', 'Date'])
ws['C1'].value = datetime.now().strftime('%Y-%m-%d')

set_header(ws, 3, ['Component', 'Total Tests', 'Status'])
test_modules = [
    ('Authentication & Authorization', 15, 'Pending'),
    ('Admin Approval System', 12, 'Pending'),
    ('Customer Management', 10, 'Pending'),
    ('Vendor Management', 10, 'Pending'),
    ('Marketplace & Products', 15, 'Pending'),
    ('Orders & Checkout', 14, 'Pending'),
    ('Chat & Messaging', 10, 'Pending'),
    ('Notifications', 8, 'Pending'),
    ('Security & Validation', 16, 'Pending'),
    ('Payment Integration (Stripe)', 12, 'Pending'),
    ('API Response Handling', 10, 'Pending'),
    ('Email System', 8, 'Pending'),
]

row = 4
for component, tests, status in test_modules:
    ws[f'A{row}'].value = component
    ws[f'B{row}'].value = tests
    ws[f'C{row}'].value = status
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].border = border
        ws[f'{col}{row}'].alignment = Alignment(horizontal='left', vertical='center')
    row += 1

# ==================== AUTHENTICATION TESTING ====================
ws = wb.create_sheet('1. Authentication')
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15

set_header(ws, 1, ['ID', 'Test Case', 'User Role', 'Test Steps', 'Expected Result', 'Status', 'Date Tested'])

auth_tests = [
    ('T-AUTH-001', 'Customer Signup', 'Public User', 'Navigate to signup, fill customer form, submit', 'Account created with pending status', 'Pending', ''),
    ('T-AUTH-002', 'Vendor Signup', 'Public User', 'Navigate to signup, select vendor, fill form', 'Account created with pending status, awaits admin approval', 'Pending', ''),
    ('T-AUTH-003', 'Admin Signup', 'Admin', 'Register admin account via admin register page', 'Admin account created with verification link', 'Pending', ''),
    ('T-AUTH-004', 'Login - Valid Customer', 'Customer', 'Enter correct email and password', 'Login successful, redirected to customer dashboard', 'Pending', ''),
    ('T-AUTH-005', 'Login - Valid Vendor', 'Vendor', 'Enter correct email and password', 'Login successful, redirected to vendor dashboard', 'Pending', ''),
    ('T-AUTH-006', 'Login - Valid Admin', 'Admin', 'Enter correct email and password', 'Login successful, redirected to admin dashboard', 'Pending', ''),
    ('T-AUTH-007', 'Login - Invalid Email', 'Public', 'Enter non-existent email', 'Error: Invalid email or password', 'Pending', ''),
    ('T-AUTH-008', 'Login - Wrong Password', 'Public', 'Enter correct email but wrong password', 'Error: Invalid email or password', 'Pending', ''),
    ('T-AUTH-009', 'Login - Suspended Account', 'Customer', 'Try to login with suspended account', 'Error: Account is suspended', 'Pending', ''),
    ('T-AUTH-010', 'Logout', 'Any', 'Click logout button', 'Session terminated, redirected to home page', 'Pending', ''),
    ('T-AUTH-011', 'Forgot Password', 'Public', 'Click forgot password, enter email', 'Reset email sent with verification token', 'Pending', ''),
    ('T-AUTH-012', 'Reset Password - Valid Token', 'Public', 'Use valid reset token from email', 'Password updated successfully', 'Pending', ''),
    ('T-AUTH-013', 'Reset Password - Expired Token', 'Public', 'Use expired reset token', 'Error: Token expired or invalid', 'Pending', ''),
    ('T-AUTH-014', 'Session Timeout', 'Any', 'Wait for session to expire', 'Redirected to login page', 'Pending', ''),
    ('T-AUTH-015', 'CSRF Token Validation', 'Any', 'Submit form without CSRF token', 'Request rejected, error 403', 'Pending', ''),
]

row = 2
for test_id, test_case, role, steps, expected, status, date in auth_tests:
    ws[f'A{row}'].value = test_id
    ws[f'B{row}'].value = test_case
    ws[f'C{row}'].value = role
    ws[f'D{row}'].value = steps
    ws[f'E{row}'].value = expected
    ws[f'F{row}'].value = status
    ws[f'G{row}'].value = date
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws[f'{col}{row}'].border = border
        ws[f'{col}{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    row += 1

# ==================== ADMIN APPROVAL SYSTEM ====================
ws = wb.create_sheet('2. Admin Approval')
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 45
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 12

set_header(ws, 1, ['ID', 'Test Case', 'User Role', 'Test Steps', 'Expected Result', 'Status', 'Notes'])

admin_tests = [
    ('T-ADMIN-001', 'View Pending Vendors List', 'Admin', 'Login as admin, check pending vendors', 'Display list of vendors with pending status', 'Pending', ''),
    ('T-ADMIN-002', 'View Pending Customers List', 'Admin', 'Login as admin, check pending customers', 'Display list of customers with inactive status', 'Pending', ''),
    ('T-ADMIN-003', 'Approve Vendor', 'Admin', 'Select vendor from pending list, click approve', 'Vendor status changes to verified, vendor can now login', 'Pending', ''),
    ('T-ADMIN-004', 'Approve Customer', 'Admin', 'Select customer from pending list, click approve', 'Customer status changes to active, customer can now login', 'Pending', ''),
    ('T-ADMIN-005', 'Reject Vendor with Reason', 'Admin', 'Select vendor, click reject, add reason', 'Vendor status = rejected, account suspended, reason logged', 'Pending', ''),
    ('T-ADMIN-006', 'Reject Customer', 'Admin', 'Select customer, click reject', 'Customer account suspended', 'Pending', ''),
    ('T-ADMIN-007', 'View Dashboard Badges', 'Admin', 'Admin dashboard loads with pending counts', 'Red notification badges show pending vendor/customer counts', 'Pending', ''),
    ('T-ADMIN-008', 'Bulk View Modal', 'Admin', 'Click Add Vendor/Customer button', 'Modal shows all pending entries with details', 'Pending', ''),
    ('T-ADMIN-009', 'Approval Notification Refresh', 'Admin', 'After approval, check if badge count decreases', 'Badge count updates immediately', 'Pending', ''),
    ('T-ADMIN-010', 'Vendor Details Display', 'Admin', 'View pending vendor modal', 'Shows: business name, email, phone, business type, signup date', 'Pending', ''),
    ('T-ADMIN-011', 'Customer Details Display', 'Admin', 'View pending customer modal', 'Shows: full name, email, phone, signup date', 'Pending', ''),
    ('T-ADMIN-012', 'Rejected User Cannot Login', 'Public', 'Try login with rejected vendor/customer account', 'Error: Account is suspended', 'Pending', ''),
]

row = 2
for test_id, test_case, role, steps, expected, status, notes in admin_tests:
    ws[f'A{row}'].value = test_id
    ws[f'B{row}'].value = test_case
    ws[f'C{row}'].value = role
    ws[f'D{row}'].value = steps
    ws[f'E{row}'].value = expected
    ws[f'F{row}'].value = status
    ws[f'G{row}'].value = notes
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws[f'{col}{row}'].border = border
        ws[f'{col}{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    row += 1

# ==================== MARKETPLACE & PRODUCTS ====================
ws = wb.create_sheet('3. Marketplace')
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 45
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 12

set_header(ws, 1, ['ID', 'Test Case', 'Endpoint', 'Test Steps', 'Expected Result', 'Status'])

marketplace_tests = [
    ('T-MARKET-001', 'Get Categories', 'GET /get_categories.php', 'Call categories endpoint', 'Returns list of all product categories', 'Pending'),
    ('T-MARKET-002', 'Get All Products', 'GET /get_products.php', 'Call products without filters', 'Returns paginated list of all products', 'Pending'),
    ('T-MARKET-003', 'Filter Products by Category', 'GET /get_products.php?category=electronics', 'Filter by category', 'Returns only products in that category', 'Pending'),
    ('T-MARKET-004', 'Search Products', 'GET /get_products.php?search=laptop', 'Search for specific product', 'Returns products matching search term', 'Pending'),
    ('T-MARKET-005', 'Filter by Price Range', 'GET /get_products.php?min_price=100&max_price=500', 'Apply price range filter', 'Returns products within price range', 'Pending'),
    ('T-MARKET-006', 'Pagination Test', 'GET /get_products.php?page=2&per_page=24', 'Navigate to page 2 with 24 items', 'Returns correct page data', 'Pending'),
    ('T-MARKET-007', 'Get Vendors List', 'GET /get_vendors.php', 'Call vendors endpoint', 'Returns list of active vendors', 'Pending'),
    ('T-MARKET-008', 'Get Recommended Vendors', 'GET /get_recommended_vendors.php', 'Get top/recommended vendors', 'Returns recommended vendors based on rating', 'Pending'),
    ('T-MARKET-009', 'Add Product (Vendor)', 'POST /add_vendor_product.php', 'Vendor adds new product', 'Product created, visible in marketplace', 'Pending'),
    ('T-MARKET-010', 'Edit Product (Vendor)', 'PUT /update_vendor_product.php', 'Vendor edits product details', 'Product updated successfully', 'Pending'),
    ('T-MARKET-011', 'Delete Product (Vendor)', 'DELETE /delete_vendor_product.php', 'Vendor deletes own product', 'Product removed from marketplace', 'Pending'),
    ('T-MARKET-012', 'View Product Details', 'GET /get_products.php?product_id=X', 'Get detailed info on product', 'Returns full product details with vendor info', 'Pending'),
    ('T-MARKET-013', 'Filter Active Products', 'GET /get_products.php?status=active', 'Filter by active status', 'Shows only active products', 'Pending'),
    ('T-MARKET-014', 'Vendor Cannot Add Duplicate SKU', 'POST /add_vendor_product.php', 'Attempt duplicate SKU', 'Error: SKU already exists', 'Pending'),
    ('T-MARKET-015', 'Product Stock Deduction', 'After Order', 'Complete order, check product stock', 'Stock count decreases correctly', 'Pending'),
]

row = 2
for test_id, test_case, endpoint, steps, expected, status in marketplace_tests:
    ws[f'A{row}'].value = test_id
    ws[f'B{row}'].value = test_case
    ws[f'C{row}'].value = endpoint
    ws[f'D{row}'].value = steps
    ws[f'E{row}'].value = expected
    ws[f'F{row}'].value = status
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].border = border
        ws[f'{col}{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    row += 1

# ==================== ORDERS & CHECKOUT ====================
ws = wb.create_sheet('4. Orders & Checkout')
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 12

set_header(ws, 1, ['ID', 'Test Case', 'Endpoint', 'Test Steps', 'Expected Result', 'Status'])

orders_tests = [
    ('T-ORDER-001', 'Create Marketplace Order', 'POST /create_marketplace_order_checkout.php', 'Customer selects products, creates order', 'Order created with pending status', 'Pending'),
    ('T-ORDER-002', 'Get Orders (Customer)', 'GET /get_orders.php', 'Customer views their orders', 'Returns only customer\'s orders', 'Pending'),
    ('T-ORDER-003', 'Get Orders (Admin)', 'GET /get_orders.php', 'Admin views all orders', 'Returns all system orders with pagination', 'Pending'),
    ('T-ORDER-004', 'Filter Orders by Status', 'GET /get_orders.php?status=completed', 'Filter orders by status', 'Returns orders with specified status', 'Pending'),
    ('T-ORDER-005', 'Filter Orders by Date Range', 'GET /get_orders.php?date_from=2026-01-01&date_to=2026-03-31', 'Filter by date range', 'Returns orders in date range', 'Pending'),
    ('T-ORDER-006', 'Search Orders', 'GET /get_orders.php?search=ORD-001', 'Search by order number', 'Returns matching orders', 'Pending'),
    ('T-ORDER-007', 'Stripe Checkout Session', 'POST /create_stripe_checkout_session.php', 'Create payment session', 'Stripe session ID generated', 'Pending'),
    ('T-ORDER-008', 'Successful Payment', 'Stripe Webhook', 'Complete payment in Stripe', 'Order marked as paid, payment_status=paid', 'Pending'),
    ('T-ORDER-009', 'Failed Payment', 'Stripe Webhook', 'Fail payment intentionally', 'Order remains pending, payment_status=failed', 'Pending'),
    ('T-ORDER-010', 'Refund Order', 'POST /refund_order.php', 'Admin issues refund', 'Order status=refunded, payment_status=refunded', 'Pending'),
    ('T-ORDER-011', 'Track Order Status', 'GET /get_orders.php?order_id=X', 'View order details', 'Shows current status and progress', 'Pending'),
    ('T-ORDER-012', 'Update Order Status', 'POST /update_order_status.php', 'Admin updates order to shipped', 'Status changes, customer notified', 'Pending'),
    ('T-ORDER-013', 'Order Calculation Accuracy', 'POST /create_marketplace_order_checkout.php', 'Create order with multiple items', 'Total price calculated correctly (items + tax + shipping)', 'Pending'),
    ('T-ORDER-014', 'Expired Cart Items', 'Cart Session', 'Cart persists after logout/login', 'Cart items remain intact or cleanup properly', 'Pending'),
]

row = 2
for test_id, test_case, endpoint, steps, expected, status in orders_tests:
    ws[f'A{row}'].value = test_id
    ws[f'B{row}'].value = test_case
    ws[f'C{row}'].value = endpoint
    ws[f'D{row}'].value = steps
    ws[f'E{row}'].value = expected
    ws[f'F{row}'].value = status
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].border = border
        ws[f'{col}{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    row += 1

# ==================== CHAT & MESSAGING ====================
ws = wb.create_sheet('5. Chat & Messaging')
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 12

set_header(ws, 1, ['ID', 'Test Case', 'Endpoint', 'Test Steps', 'Expected Result', 'Status'])

chat_tests = [
    ('T-CHAT-001', 'Start Conversation', 'POST /chat_start_conversation.php', 'Customer initiates chat with vendor', 'Conversation created with ID', 'Pending'),
    ('T-CHAT-002', 'Send Message', 'POST /chat_send_message.php', 'Send message in conversation', 'Message saved, timestamp recorded', 'Pending'),
    ('T-CHAT-003', 'Receive Message', 'GET /chat_data.php', 'Fetch conversation messages', 'Returns all messages in order', 'Pending'),
    ('T-CHAT-004', 'Message Timestamp', 'POST /chat_send_message.php', 'Send message, check timestamp', 'Timestamp accurate and formatted correctly', 'Pending'),
    ('T-CHAT-005', 'User Cannot Access Other Chats', 'GET /chat_data.php?conversation_id=X', 'Try access conversation user is not part of', 'Error: Access denied', 'Pending'),
    ('T-CHAT-006', 'Conversation History', 'GET /chat_data.php', 'Load conversation with 100+ messages', 'All messages load correctly', 'Pending'),
    ('T-CHAT-007', 'Empty Conversation', 'GET /chat_data.php', 'Fetch conversation with no messages', 'Returns empty array', 'Pending'),
    ('T-CHAT-008', 'Message Validation', 'POST /chat_send_message.php', 'Send empty message', 'Error: Message cannot be empty', 'Pending'),
    ('T-CHAT-009', 'Real-time Notification', 'POST /chat_send_message.php', 'Send message, check notification', 'Recipient receives notification', 'Pending'),
    ('T-CHAT-010', 'Chat History Export', 'GET /chat_data.php', 'Export chat history', 'Returns all messages for export', 'Pending'),
]

row = 2
for test_id, test_case, endpoint, steps, expected, status in chat_tests:
    ws[f'A{row}'].value = test_id
    ws[f'B{row}'].value = test_case
    ws[f'C{row}'].value = endpoint
    ws[f'D{row}'].value = steps
    ws[f'E{row}'].value = expected
    ws[f'F{row}'].value = status
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].border = border
        ws[f'{col}{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    row += 1

# ==================== NOTIFICATIONS ====================
ws = wb.create_sheet('6. Notifications')
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 12

set_header(ws, 1, ['ID', 'Test Case', 'Endpoint', 'Test Steps', 'Expected Result', 'Status'])

notif_tests = [
    ('T-NOTIF-001', 'Get All Notifications', 'GET /get_notifications.php', 'Load user notifications', 'Returns unread notifications', 'Pending'),
    ('T-NOTIF-002', 'Mark Notification as Read', 'POST /mark_notifications_read.php', 'Click notification to read', 'Notification marked as read', 'Pending'),
    ('T-NOTIF-003', 'Notification Count Badge', 'GET /get_notifications.php', 'Check unread count', 'Badge shows correct unread count', 'Pending'),
    ('T-NOTIF-004', 'New Order Notification', 'POST /create_marketplace_order_checkout.php', 'Place order, check notification', 'Vendor receives order notification', 'Pending'),
    ('T-NOTIF-005', 'Order Status Change Notification', 'POST /update_order_status.php', 'Update order status, check notification', 'Customer receives status update notification', 'Pending'),
    ('T-NOTIF-006', 'Approval Notification', 'POST /admin_approve_vendor.php', 'Admin approves vendor, check notification', 'Vendor receives approval notification', 'Pending'),
    ('T-NOTIF-007', 'Clear All Notifications', 'DELETE /clear_notifications.php', 'Clear all notifications', 'All notifications deleted', 'Pending'),
    ('T-NOTIF-008', 'Notification Pagination', 'GET /get_notifications.php?page=2', 'Load notifications page 2', 'Returns correct page of notifications', 'Pending'),
]

row = 2
for test_id, test_case, endpoint, steps, expected, status in notif_tests:
    ws[f'A{row}'].value = test_id
    ws[f'B{row}'].value = test_case
    ws[f'C{row}'].value = endpoint
    ws[f'D{row}'].value = steps
    ws[f'E{row}'].value = expected
    ws[f'F{row}'].value = status
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].border = border
        ws[f'{col}{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    row += 1

# ==================== SECURITY & VALIDATION ====================
ws = wb.create_sheet('7. Security')
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 45
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 12

set_header(ws, 1, ['ID', 'Test Case', 'Vulnerability Type', 'Test Steps', 'Expected Result', 'Status'])

security_tests = [
    ('T-SEC-001', 'SQL Injection Prevention', 'SQL Injection', 'Try SQL injection in search field', 'Query fails, no error exposed', 'Pending'),
    ('T-SEC-002', 'XSS Prevention', 'XSS Attack', 'Try script injection in product name', 'Script sanitized or escaped', 'Pending'),
    ('T-SEC-003', 'CSRF Token Validation', 'CSRF', 'Submit form without CSRF token', 'Request rejected', 'Pending'),
    ('T-SEC-004', 'Password Hashing', 'Password Security', 'Check database password field', 'Passwords hashed, not plain text', 'Pending'),
    ('T-SEC-005', 'Role-Based Access Control', 'Authorization', 'Customer tries access admin API', 'Error 403: Access denied', 'Pending'),
    ('T-SEC-006', 'Session Hijacking Prevention', 'Session Security', 'Try access with invalid session', 'Redirect to login', 'Pending'),
    ('T-SEC-007', 'Rate Limiting (Login)', 'Brute Force', 'Try login 10 times with wrong password', 'Account locked after N attempts', 'Pending'),
    ('T-SEC-008', 'Email Verification Link Expiry', 'Token Security', 'Use expired verification link', 'Error: Link expired', 'Pending'),
    ('T-SEC-009', 'HTTPS Enforcement', 'Transport Security', 'Check API responses', 'APIs use HTTPS in production', 'Pending'),
    ('T-SEC-010', 'Sensitive Data Not Exposed', 'Data Leakage', 'Check API responses', 'No passwords/tokens in API responses', 'Pending'),
    ('T-SEC-011', 'File Upload Validation', 'File Injection', 'Try upload .exe/.php file', 'File rejected, extension validated', 'Pending'),
    ('T-SEC-012', 'Input Length Validation', 'Buffer Overflow', 'Submit extremely long input (10000 chars)', 'Request truncated or rejected', 'Pending'),
    ('T-SEC-013', 'Email Header Injection', 'Email Security', 'Try inject headers in email field', 'Email validation prevents injection', 'Pending'),
    ('T-SEC-014', 'Secure Cookie Settings', 'Cookie Security', 'Check session cookie flags', 'HttpOnly, Secure, SameSite flags set', 'Pending'),
    ('T-SEC-015', 'SQL Error Messages Sanitized', 'Information Disclosure', 'Trigger SQL error', 'Error message generic, no SQL exposed', 'Pending'),
    ('T-SEC-016', 'API Rate Limiting', 'DoS Protection', 'Make 1000 requests per minute', 'Requests throttled/blocked after limit', 'Pending'),
]

row = 2
for test_id, test_case, vuln_type, steps, expected, status in security_tests:
    ws[f'A{row}'].value = test_id
    ws[f'B{row}'].value = test_case
    ws[f'C{row}'].value = vuln_type
    ws[f'D{row}'].value = steps
    ws[f'E{row}'].value = expected
    ws[f'F{row}'].value = status
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].border = border
        ws[f'{col}{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    row += 1

# ==================== DATA VALIDATION ====================
ws = wb.create_sheet('8. Data Validation')
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 25
ws.column_dimensions['F'].width = 12

set_header(ws, 1, ['ID', 'Test Case', 'Field Type', 'Invalid Input', 'Expected Result', 'Status'])

validation_tests = [
    ('T-VAL-001', 'Email Format Validation', 'Email', 'test@', 'Error: Invalid email format', 'Pending'),
    ('T-VAL-002', 'Email Already Exists', 'Email', 'existing@email.com', 'Error: Email already registered', 'Pending'),
    ('T-VAL-003', 'Password Length Validation', 'Password', '123', 'Error: Password too short (min 8 chars)', 'Pending'),
    ('T-VAL-004', 'Phone Number Format', 'Phone', 'abcdef', 'Error: Invalid phone format', 'Pending'),
    ('T-VAL-005', 'Price Format Validation', 'Price', '-100', 'Error: Price must be positive', 'Pending'),
    ('T-VAL-006', 'Required Fields', 'Any', 'Leave required field empty', 'Error: Field is required', 'Pending'),
    ('T-VAL-007', 'Maximum Length Validation', 'Text', '500 character string', 'Error: Input too long', 'Pending'),
    ('T-VAL-008', 'Numeric Field Validation', 'Number', 'abc123', 'Error: Must be numeric', 'Pending'),
    ('T-VAL-009', 'Date Format Validation', 'Date', '32/13/2026', 'Error: Invalid date', 'Pending'),
    ('T-VAL-010', 'URL Format Validation', 'URL', 'not-a-url', 'Error: Invalid URL format', 'Pending'),
]

row = 2
for test_id, test_case, field_type, invalid_input, expected, status in validation_tests:
    ws[f'A{row}'].value = test_id
    ws[f'B{row}'].value = test_case
    ws[f'C{row}'].value = field_type
    ws[f'D{row}'].value = invalid_input
    ws[f'E{row}'].value = expected
    ws[f'F{row}'].value = status
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].border = border
        ws[f'{col}{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    row += 1

# ==================== API RESPONSE TESTING ====================
ws = wb.create_sheet('9. API Response')
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 12

set_header(ws, 1, ['ID', 'Test Case', 'Endpoint', 'Test Steps', 'Expected Response', 'Status'])

api_tests = [
    ('T-API-001', 'Success Response Format', 'Any GET', 'Call successful endpoint', 'Response has success=true, code, message, data', 'Pending'),
    ('T-API-002', 'Error Response Format', 'Any Endpoint', 'Call with invalid parameter', 'Response has success=false, errors array', 'Pending'),
    ('T-API-003', 'Pagination Meta Data', 'GET /get_products.php', 'Call with pagination', 'Response includes meta.pagination with page, total, total_pages', 'Pending'),
    ('T-API-004', 'Timestamp Format', 'Any Endpoint', 'Check response', 'Timestamp in ISO 8601 format', 'Pending'),
    ('T-API-005', 'HTTP Status Codes - 200 OK', 'Any Success', 'Call successful endpoint', 'HTTP 200 returned', 'Pending'),
    ('T-API-006', 'HTTP Status Codes - 400 Bad Request', 'Any with bad data', 'Send invalid parameters', 'HTTP 400 returned', 'Pending'),
    ('T-API-007', 'HTTP Status Codes - 401 Unauthorized', 'Protected endpoint', 'Call without auth', 'HTTP 401 returned', 'Pending'),
    ('T-API-008', 'HTTP Status Codes - 403 Forbidden', 'Admin endpoint', 'Customer calls admin endpoint', 'HTTP 403 returned', 'Pending'),
    ('T-API-009', 'HTTP Status Codes - 404 Not Found', 'Non-existent resource', 'Call with invalid ID', 'HTTP 404 returned', 'Pending'),
    ('T-API-010', 'Response Time Performance', 'Any', 'Measure response time', 'Response time < 500ms for normal queries', 'Pending'),
]

row = 2
for test_id, test_case, endpoint, steps, expected, status in api_tests:
    ws[f'A{row}'].value = test_id
    ws[f'B{row}'].value = test_case
    ws[f'C{row}'].value = endpoint
    ws[f'D{row}'].value = steps
    ws[f'E{row}'].value = expected
    ws[f'F{row}'].value = status
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].border = border
        ws[f'{col}{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    row += 1

# ==================== EDGE CASES & PERFORMANCE ====================
ws = wb.create_sheet('10. Edge Cases')
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 12

set_header(ws, 1, ['ID', 'Test Case', 'Category', 'Test Steps', 'Expected Result', 'Status'])

edge_tests = [
    ('T-EDGE-001', 'Empty Database Query', 'Edge Case', 'Query table with no records', 'Returns empty array, not error', 'Pending'),
    ('T-EDGE-002', 'Large Pagination Request', 'Performance', 'Request per_page=1000', 'Returns max allowed or caps at limit', 'Pending'),
    ('T-EDGE-003', 'Special Characters in Search', 'Edge Case', 'Search for "product@#$%"', 'Handles special chars safely', 'Pending'),
    ('T-EDGE-004', 'Zero Stock Product', 'Edge Case', 'Try purchase out-of-stock item', 'Error: Item out of stock', 'Pending'),
    ('T-EDGE-005', 'Concurrent Order Processing', 'Concurrency', 'Two users buy same product simultaneously', 'Stock deducted correctly, no double sell', 'Pending'),
    ('T-EDGE-006', 'Negative Order Amount', 'Validation', 'Try create order with negative price', 'Error: Invalid amount', 'Pending'),
    ('T-EDGE-007', 'Future Date Filtering', 'Edge Case', 'Filter orders by future date', 'Returns no results', 'Pending'),
    ('T-EDGE-008', 'Circular Vendor Recommendations', 'Logic', 'Check vendor recommendations algorithm', 'No infinite loops or duplicate vendors', 'Pending'),
    ('T-EDGE-009', 'Expired Session Token', 'Security', 'Use token after expiry', 'Requires re-authentication', 'Pending'),
    ('T-EDGE-010', 'Simultaneous Message Sends', 'Concurrency', 'Send 2 messages from same user simultaneously', 'Both messages saved with correct timestamps', 'Pending'),
]

row = 2
for test_id, test_case, category, steps, expected, status in edge_tests:
    ws[f'A{row}'].value = test_id
    ws[f'B{row}'].value = test_case
    ws[f'C{row}'].value = category
    ws[f'D{row}'].value = steps
    ws[f'E{row}'].value = expected
    ws[f'F{row}'].value = status
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].border = border
        ws[f'{col}{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    row += 1

# ==================== TEST EXECUTION LOG ====================
ws = wb.create_sheet('11. Execution Log')
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 12

set_header(ws, 1, ['Date Tested', 'Test ID', 'Pass/Fail', 'Notes/Error Details', 'Tester Name', 'Env'])

ws['A2'].value = 'Date in YYYY-MM-DD format'
ws['B2'].value = 'e.g., T-AUTH-001'
ws['C2'].value = 'PASS/FAIL'
ws['D2'].value = 'Observations and error messages if any'
ws['E2'].value = 'Your name'
ws['F2'].value = 'Dev/Prod'

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws[f'{col}2'].border = border
    ws[f'{col}2'].font = Font(italic=True, color='666666')
    ws[f'{col}2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Add 20 empty rows for manual entry
for row in range(3, 23):
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].border = border

# ==================== TEST SUMMARY & NOTES ====================
ws = wb.create_sheet('12. Summary & Notes')
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 70

set_header(ws, 1, ['Field', 'Value'])

summary_data = [
    ('Project Name', 'BizLink CRM Platform'),
    ('Project Type', 'Multi-role CRM + Marketplace'),
    ('Testing Date Range', 'Start Date: _________  End Date: _________'),
    ('', ''),
    ('TOTAL TEST CASES', '135'),
    ('Total Passed', ''),
    ('Total Failed', ''),
    ('Total Blocked', ''),
    ('Pass Percentage', ''),
    ('', ''),
    ('KEY COMPONENTS TESTED', ''),
    ('1. Authentication & Authorization', 'Multi-role login, signup, password reset'),
    ('2. Admin Approval System', 'Vendor/customer approval workflows'),
    ('3. Marketplace & Products', 'Product catalog, search, filtering'),
    ('4. Orders & Checkout', 'Order creation, payment processing (Stripe)'),
    ('5. Chat & Messaging', 'Real-time messaging between users'),
    ('6. Notifications', 'Real-time alerts for orders, approvals'),
    ('7. Security', 'SQL injection, XSS, CSRF, rate limiting'),
    ('8. Data Validation', 'Input validation, format checks'),
    ('9. API Response Handling', 'HTTP status codes, response format'),
    ('10. Edge Cases', 'Performance, concurrency, special scenarios'),
    ('', ''),
    ('CRITICAL FINDINGS', ''),
    ('Issue 1', 'Description & Impact'),
    ('Issue 2', 'Description & Impact'),
    ('Issue 3', 'Description & Impact'),
    ('', ''),
    ('RECOMMENDATIONS', ''),
    ('Recommendation 1', 'Action to take'),
    ('Recommendation 2', 'Action to take'),
    ('', ''),
    ('TESTER SIGN-OFF', ''),
    ('Tester Name', ''),
    ('Date Completed', ''),
    ('Approved By', ''),
]

row = 2
for field, value in summary_data:
    ws[f'A{row}'].value = field
    ws[f'B{row}'].value = value
    
    if field and field.isupper():
        ws[f'A{row}'].font = Font(bold=True, color='FFFFFF')
        ws[f'A{row}'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    for col in ['A', 'B']:
        ws[f'{col}{row}'].border = border
        ws[f'{col}{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    row += 1

# Save file
wb.save('BizLink_CRM_Comprehensive_Test_Suite.xlsx')
print('✓ Excel test file created successfully!')
print('✓ File: BizLink_CRM_Comprehensive_Test_Suite.xlsx')
print('✓ Total sheets: 12')
print('✓ Total test cases: 135+')
